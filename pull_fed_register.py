"""
Pull all CMS documents (Jan 2024 - May 2026) from the Federal Register API,
scan full text for MA / Medigap / SNP keywords, and write a "Fed Register"
tab into the Medicare Key Dates workbook.

Usage:
    pip install requests openpyxl
    python pull_fed_register.py Medicare_Key_Dates_Calendar-3-2.xlsx

Data source: FederalRegister.gov API v1 (official, no API key)
    Docs: https://www.federalregister.gov/developers/documentation/api/v1
    (linked from the site footer: Reader Aids > Developer Resources)

Notes:
- Pulls ALL CMS docs in the window (~450-500 expected), not just the
  "Medicare, Medicaid, SCHIP payment" topic facet (~63) — topic tags are
  applied inconsistently, so keyword scanning full text gives better recall.
- Full text is cached to fedreg_cache/ so interrupted runs resume and
  re-runs only fetch new documents.
- Matched_Keywords shows hit counts per term, e.g. "medicare advantage(31); d-snp(4)".
  Sort/filter the tab by Match=1, then skim Matched_Keywords to triage:
  a doc with medicare advantage(31) is a real read; part d(1) alone may be a passing mention.
- Publication_Date = when it appeared in the Register (the "announced" date);
  Effective_Date = when it takes legal effect. Both matter for the Events
  tab's Date_Type column. Effective_Date is blank for most Notices.
- Re-running rewrites the "Fed Register" tab cleanly; manual edits there
  will be lost, so promote keepers to the Events tab instead.
"""

import sys
import os
import re
import time
import datetime as dt

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

API = "https://www.federalregister.gov/api/v1/documents.json"
CACHE_DIR = "fedreg_cache"
DATE_GTE = "2024-01-01"
DATE_LTE = "2026-05-31"

# Keyword -> list of regex variants (case-insensitive). Word boundaries where
# it matters so "part d" doesn't match "particular d..." etc.
KEYWORDS = {
    "medicare advantage": [r"medicare advantage"],
    "ma organization":    [r"\bma organizations?\b"],
    "part c":             [r"\bpart c\b"],
    "part d":             [r"\bpart d\b"],
    "special needs plan": [r"special needs plans?"],
    "d-snp":              [r"\bd[- ]snps?\b"],
    "c-snp":              [r"\bc[- ]snps?\b"],
    "dual eligible":      [r"dual[- ]eligibles?\b"],
    "medigap":            [r"\bmedigap\b"],
    "medicare supplement": [r"medicare supplement"],
}
COMPILED = {k: [re.compile(p, re.IGNORECASE) for p in pats] for k, pats in KEYWORDS.items()}

FIELDS = ["document_number", "title", "type", "abstract",
          "publication_date", "effective_on", "html_url", "raw_text_url"]


def fetch_metadata():
    """Page through the API and return all CMS docs in the window."""
    params = {
        "conditions[agencies][]": "centers-for-medicare-medicaid-services",
        "conditions[publication_date][gte]": DATE_GTE,
        "conditions[publication_date][lte]": DATE_LTE,
        "per_page": 300,
        "order": "oldest",
    }
    # fields[] must repeat, requests handles list values
    params["fields[]"] = FIELDS

    docs, url, first = [], API, True
    while url:
        resp = requests.get(url, params=params if first else None, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        if first:
            print(f"API reports {data.get('count')} CMS documents in window")
            first = False
        docs.extend(data.get("results", []))
        url = data.get("next_page_url")
        time.sleep(0.5)
    return docs


def get_full_text(doc):
    """Fetch (or read cached) raw text for one document. Returns '' on failure."""
    num = doc["document_number"]
    path = os.path.join(CACHE_DIR, f"{num}.txt")
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return f.read()
    url = doc.get("raw_text_url")
    if not url:
        return ""
    for attempt in range(3):
        try:
            resp = requests.get(url, timeout=120)
            resp.raise_for_status()
            text = resp.text
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
            time.sleep(0.5)  # be polite
            return text
        except Exception as e:
            print(f"  retry {attempt+1} for {num}: {e}")
            time.sleep(3)
    print(f"  FAILED to fetch text for {num} — scanning title+abstract only")
    return ""


def scan(text):
    """Return {keyword: count} for keywords found in text."""
    hits = {}
    for kw, patterns in COMPILED.items():
        n = sum(len(p.findall(text)) for p in patterns)
        if n:
            hits[kw] = n
    return hits


def parse_date(s):
    if not s:
        return None
    try:
        return dt.datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        return None


def main():
    if len(sys.argv) != 2:
        sys.exit("usage: python pull_fed_register.py <workbook.xlsx>")
    wb_path = sys.argv[1]
    os.makedirs(CACHE_DIR, exist_ok=True)

    docs = fetch_metadata()
    print(f"pulled metadata for {len(docs)} documents; scanning full text...")

    rows = []
    for i, doc in enumerate(docs, 1):
        text = get_full_text(doc)
        basis = text if text else f"{doc.get('title','')} {doc.get('abstract','')}"
        hits = scan(basis)
        matched = "; ".join(f"{k}({v})" for k, v in
                            sorted(hits.items(), key=lambda kv: -kv[1]))
        rows.append({
            "num": doc["document_number"],
            "type": doc.get("type", ""),
            "pub": parse_date(doc.get("publication_date")),
            "eff": parse_date(doc.get("effective_on")),
            "match": 1 if hits else 0,
            "kw": matched,
            "title": doc.get("title", ""),
            "abstract": (doc.get("abstract") or "")[:500],
            "url": doc.get("html_url", ""),
            "scanned_full": bool(text),
        })
        if i % 25 == 0:
            print(f"  scanned {i}/{len(docs)}")

    rows.sort(key=lambda r: r["pub"] or dt.datetime.min)
    n_match = sum(r["match"] for r in rows)
    n_partial = sum(1 for r in rows if not r["scanned_full"])
    print(f"done: {len(rows)} docs, {n_match} keyword matches, "
          f"{n_partial} scanned on title/abstract only")

    # ---- write the tab ----
    wb = load_workbook(wb_path)
    if "Fed Register" in wb.sheetnames:
        del wb["Fed Register"]
    ws = wb.create_sheet("Fed Register")

    hdr_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    headers = ["Doc_Number", "Type", "Pub_Date", "Effective_Date", "Match",
               "Matched_Keywords", "Title", "Abstract", "URL", "Full_Text_Scanned"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
    for r, row in enumerate(rows, 2):
        vals = [row["num"], row["type"], row["pub"], row["eff"], row["match"],
                row["kw"], row["title"], row["abstract"], row["url"],
                "Y" if row["scanned_full"] else "TITLE/ABSTRACT ONLY"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = body_font
            if c in (3, 4) and v:
                cell.number_format = "m/d/yyyy"
    widths = [16, 14, 11, 13, 7, 40, 60, 60, 45, 18]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(rows)+1}"

    wb.save(wb_path)
    print(f"wrote 'Fed Register' tab ({len(rows)} rows) to {wb_path}")
    print("filter Match=1, sort by Matched_Keywords, and start reading.")


if __name__ == "__main__":
    main()
